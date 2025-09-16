import csv
import io
from typing import List, Optional

from django.db import transaction
from django.http import HttpResponse
from django.db.models import Q
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.decorators import api_view

from .models import Lesson, Room
from apps.courses.models import Course
from apps.teachers.models import Teacher
from apps.schools.models import Class as SchoolClass
from .serializers import LessonSerializer


def try_decode(data: bytes) -> str:
    for enc in ['utf-8-sig','utf-8','gbk','gb18030','cp936']:
        try:
            return data.decode(enc)
        except Exception:
            continue
    raise UnicodeDecodeError('unknown','',0,1,'无法解码文件，请使用 UTF-8 或 GBK')


class TimetableImportView(APIView):
    @transaction.atomic
    def post(self, request):
        file = request.FILES.get('file')
        term = request.POST.get('term')
        mode = request.POST.get('mode','append')
        if not file or not term:
            return Response({
                'success': False,
                'error': {'code':'VALIDATION_ERROR','message':'文件与学期必填','details':{'file':['必填'],'term':['必填']}},
            }, status=status.HTTP_400_BAD_REQUEST)

        if mode == 'overwrite':
            Lesson.objects.filter(term=term).delete()

        # 若为 Excel，优先按 Excel 表格课表样式解析；否则回退 CSV
        name = getattr(file, 'name', '').lower()
        rows = []
        if name.endswith('.xlsx') or name.endswith('.xls'):
            try:
                import openpyxl  # type: ignore
                wb = openpyxl.load_workbook(file, data_only=True)
                ws = wb.active
                # 支持两种 Excel 样式：
                # A) 第一列是“班级”，第一行是“星期*”合并，第二行“上午/下午”，第三行 1..9 节；第4行开始为各班级
                # B) 第一列是“节次/时间”，第一行是“星期*”，第二行可选“上午/下午”，后续为时间行

                A1 = str(ws.cell(row=1, column=1).value or '').strip()
                # 通用：解析第一行的星期列映射
                headers = []
                for col in ws.iter_cols(min_row=1, max_row=1):
                    headers.append(str(col[0].value or '').strip())
                day_alias = {'周一':1,'星期一':1,'一':1,'周二':2,'星期二':2,'二':2,'周三':3,'星期三':3,'三':3,'周四':4,'星期四':4,'四':4,'周五':5,'星期五':5,'五':5,'周六':6,'星期六':6,'六':6,'周日':7,'星期日':7,'日':7}
                # 处理合并单元格：向右填充最后出现的星期
                day_cols = []
                last_d = None
                for h in headers:
                    d = day_alias.get(h, None)
                    if d:
                        last_d = d
                    day_cols.append(d or last_d)
                # 若整行为空，尝试用工作表标题判断
                day_from_ws_title = day_alias.get(ws.title, None)

                if '班级' in A1:
                    # 样式A：按班级为行
                    # 第三行应为节次数字；若不是，尝试从第二行读取
                    period_row = 3
                    period_values = [str(c.value or '').strip() for c in ws[period_row]] if ws.max_row >= 3 else []
                    if not any(v == '1' for v in period_values):
                        period_row = 2
                        period_values = [str(c.value or '').strip() for c in ws[period_row]]
                    start_row = 4 if period_row == 3 else 3
                    max_cols = ws.max_column
                    # 遍历每个班级行
                    for r in range(start_row, ws.max_row + 1):
                        class_name = str(ws.cell(row=r, column=1).value or '').strip()
                        if not class_name:
                            continue
                        for c_idx in range(2, max_cols + 1):
                            day = day_cols[c_idx - 1] if c_idx - 1 < len(day_cols) else None
                            if not day:
                                day = day_from_ws_title
                            if not day:
                                continue
                            p_val = str(ws.cell(row=period_row, column=c_idx).value or '').strip()
                            if not p_val.isdigit():
                                continue
                            start_period = int(p_val)
                            cell_val = ws.cell(row=r, column=c_idx).value
                            if not cell_val:
                                continue
                            text = str(cell_val).strip()
                            if not text:
                                continue
                            course_name = text.split('\n')[0].strip()
                            rest = text.split('\n')[1].strip() if '\n' in text else ''
                            teacher_name = ''
                            room_name = ''
                            if rest:
                                parts = [p.strip() for p in rest.replace('＠','@').split('@')]
                                teacher_name = parts[0] if parts else ''
                                room_name = parts[1] if len(parts) > 1 else ''
                            rows.append({
                                'day': day,
                                'course_name': course_name,
                                'teacher_name': teacher_name,
                                'room': room_name,
                                'class_name': class_name,
                                'start_period': start_period,
                                'end_period': start_period,
                            })
                else:
                    # 样式B：按时间为行（旧方案）
                    second_row_values = [str(c.value or '').strip() for c in ws[2]] if ws.max_row >= 2 else []
                    has_ampm = any(v in ['上午','下午'] for v in second_row_values)
                    start_row = 3 if has_ampm else 2
                    for r in ws.iter_rows(min_row=start_row):
                        time_label = str(r[0].value or '').strip()
                        for c_idx in range(1, len(r)):
                            day = day_cols[c_idx] if c_idx < len(day_cols) else None
                            if not day:
                                day = day_from_ws_title
                            if not day:
                                continue
                            cell = r[c_idx].value
                            if not cell:
                                continue
                            text = str(cell).strip()
                            if not text:
                                continue
                            course_name = text.split('\n')[0].strip()
                            rest = text.split('\n')[1].strip() if '\n' in text else ''
                            teacher_name = ''
                            room_name = ''
                            if rest:
                                parts = [p.strip() for p in rest.replace('＠','@').split('@')]
                                teacher_name = parts[0] if parts else ''
                                room_name = parts[1] if len(parts) > 1 else ''
                            rows.append({
                                'day': day,
                                'course_name': course_name,
                                'teacher_name': teacher_name,
                                'room': room_name,
                                'time_label': time_label,
                            })
            except Exception as e:
                return Response({'success': False, 'error': {'code':'INVALID_FILE', 'message': f'Excel 解析失败: {e}'}}, status=status.HTTP_400_BAD_REQUEST)
        else:
            content = file.read()
            text = try_decode(content)
            f = io.StringIO(text)
            reader = csv.DictReader(f)
            rows = list(reader)
        normalized_rows: List[dict] = []

        # 表头同义映射
        alias = {
            'term':['term','学期'],
            'weeks':['weeks','周次'],
            'day':['day','星期','星期几'],
            'start_time':['start_time','开始时间'],
            'end_time':['end_time','结束时间'],
            'start_period':['start_period','开始节次'],
            'end_period':['end_period','结束节次'],
            'course_name':['course_name','课程名称','课程'],
            'teacher_name':['teacher_name','教师名称','老师','授课老师'],
            'class_name':['class_name','班级名称','班级'],
            'room':['room','教室','教室名称'],
            'week_type':['week_type','单双周'],
            'remark':['remark','备注'],
        }

        def get_val(row, key):
            for k in alias.get(key, [key]):
                if k in row and row[k] is not None:
                    return str(row[k]).strip()
            return ''

        day_map = {'一':1,'二':2,'三':3,'四':4,'五':5,'六':6,'日':7,'天':7,'周一':1,'周二':2,'周三':3,'周四':4,'周五':5,'周六':6,'周日':7,'Mon':1,'Tue':2,'Wed':3,'Thu':4,'Fri':5,'Sat':6,'Sun':7}

        created = 0
        for row in rows:
            course_name = get_val(row,'course_name')
            if not course_name:
                continue
            d = str(get_val(row,'day'))
            if d.isdigit():
                day_of_week = int(d)
            else:
                day_of_week = day_map.get(d, 1)
            # Excel 课表样式下可用 time_label 推断时间/节次；这里先直接落时间为空，由前端按节次渲染
            start_time = get_val(row,'start_time')
            end_time = get_val(row,'end_time')
            start_period = get_val(row,'start_period')
            end_period = get_val(row,'end_period')
            teacher_name = get_val(row,'teacher_name')
            class_name = get_val(row,'class_name')
            room_name = get_val(row,'room')
            week_type = get_val(row,'week_type')
            weeks = get_val(row,'weeks')
            remark = get_val(row,'remark')

            course = Course.objects.filter(name=course_name).first()
            teacher = Teacher.objects.filter(name=teacher_name).first() if teacher_name else None
            class_obj = SchoolClass.objects.filter(name=class_name).first() if class_name else None
            room = Room.objects.filter(name=room_name).first() if room_name else None
            if room_name and not room:
                room = Room.objects.create(name=room_name)

            Lesson.objects.create(
                term=term,
                day_of_week=day_of_week,
                start_time=start_time or None,
                end_time=end_time or None,
                start_period=int(start_period) if str(start_period).isdigit() else None,
                end_period=int(end_period) if str(end_period).isdigit() else None,
                week_type='odd' if week_type == '单' else ('even' if week_type == '双' else 'all'),
                weeks=weeks,
                course=course,
                teacher=teacher,
                class_ref=class_obj,
                room=room,
                course_name=course_name,
                teacher_name=teacher_name,
                class_name=class_name,
                room_name=room_name,
                remark=remark,
            )
            created += 1

        return Response({'success': True, 'data': {'created': created}})


@api_view(['POST'])
def create_lesson(request):
    """手动创建单条课次（Lesson）。
    支持两种外键入参：优先使用 *_id；若无则按 *_name 匹配（room 不存在则自动创建）。
    weeks 可为字符串（如 "1-16" 或 "1,3,5"）或 number[]（将拼接为逗号分隔）。
    weekType 支持 'odd'|'even'|'all' 或 中文 '单'|'双'。
    """
    data = request.data

    term = (data.get('term') or '').strip()
    day = data.get('dayOfWeek') or data.get('day_of_week')
    course_name = (data.get('courseName') or data.get('course_name') or '').strip()
    if not term or not str(day).strip() or not course_name:
        return Response({
            'success': False,
            'error': {
                'code': 'VALIDATION_ERROR',
                'message': 'term、dayOfWeek、courseName 为必填',
                'details': {
                    'term': ['必填' if not term else None],
                    'dayOfWeek': ['必填' if not str(day).strip() else None],
                    'courseName': ['必填' if not course_name else None],
                }
            }
        }, status=status.HTTP_400_BAD_REQUEST)

    # 规范化 weekType
    week_type_raw = (data.get('weekType') or data.get('week_type') or 'all').strip()
    if week_type_raw in ['单', 'odd']:
        week_type = 'odd'
    elif week_type_raw in ['双', 'even']:
        week_type = 'even'
    else:
        week_type = 'all'

    # 规范化 weeks
    weeks_val = data.get('weeks')
    if isinstance(weeks_val, list):
        weeks = ','.join(str(int(x)) for x in weeks_val if str(x).strip())
    else:
        weeks = str(weeks_val or '').strip()

    # 外键优先使用 *_id
    course = None
    course_id = data.get('courseId') or data.get('course_id')
    if course_id:
        course = Course.objects.filter(pk=course_id).first()
    elif course_name:
        course = Course.objects.filter(name=course_name).first()

    teacher = None
    teacher_id = data.get('teacherId') or data.get('teacher_id')
    teacher_name = (data.get('teacherName') or data.get('teacher_name') or '').strip()
    if teacher_id:
        teacher = Teacher.objects.filter(pk=teacher_id).first()
        # 如果找到了教师对象，使用其名称
        if teacher and not teacher_name:
            teacher_name = teacher.name
    elif teacher_name:
        teacher = Teacher.objects.filter(name=teacher_name).first()

    class_obj = None
    class_id = data.get('classId') or data.get('class_id') or data.get('class_ref_id')
    class_name = (data.get('className') or data.get('class_name') or '').strip()
    if class_id:
        class_obj = SchoolClass.objects.filter(pk=class_id).first()
        # 如果找到了班级对象，使用其名称
        if class_obj and not class_name:
            class_name = class_obj.name
    elif class_name:
        class_obj = SchoolClass.objects.filter(name=class_name).first()

    room = None
    room_id = data.get('roomId') or data.get('room_id')
    room_name = (data.get('roomName') or data.get('room_name') or '').strip()
    if room_id:
        room = Room.objects.filter(pk=room_id).first()
        # 如果找到了教室对象，使用其名称
        if room and not room_name:
            room_name = room.name
    elif room_name:
        room = Room.objects.filter(name=room_name).first()
        if not room:
            room = Room.objects.create(name=room_name)

    payload = {
        'term': term,
        'dayOfWeek': int(day),
        'weekType': week_type,
        'courseName': course_name,
    }
    # 可选字段按有值再放入，避免空字符串触发校验错误
    start_time_val = (data.get('startTime') or data.get('start_time') or '').strip()
    end_time_val = (data.get('endTime') or data.get('end_time') or '').strip()
    if start_time_val:
        payload['startTime'] = start_time_val
    if end_time_val:
        payload['endTime'] = end_time_val
    start_period_val = data.get('startPeriod') or data.get('start_period')
    end_period_val = data.get('endPeriod') or data.get('end_period')
    if start_period_val is not None and str(start_period_val) != '':
        payload['startPeriod'] = start_period_val
    if end_period_val is not None and str(end_period_val) != '':
        payload['endPeriod'] = end_period_val
    if weeks:
        payload['weeks'] = weeks
    cid = str(course.id) if course else (course_id or None)
    if cid:
        payload['courseId'] = cid
    tid = str(teacher.id) if teacher else (teacher_id or None)
    if tid:
        payload['teacherId'] = tid
    if teacher_name:
        payload['teacherName'] = teacher_name
    clsid = str(class_obj.id) if class_obj else (class_id or None)
    if clsid:
        payload['classId'] = clsid
    if class_name:
        payload['className'] = class_name
    rid = str(room.id) if room else (room_id or None)
    if rid:
        payload['roomId'] = rid
    if room_name:
        payload['roomName'] = room_name
    remark_val = (data.get('remark') or '').strip()
    if remark_val:
        payload['remark'] = remark_val

    # 序列化与保存
    ser = LessonSerializer(data=payload)
    if not ser.is_valid():
        return Response({'success': False, 'error': {'code': 'VALIDATION_ERROR', 'message': '数据验证失败', 'details': ser.errors}}, status=status.HTTP_400_BAD_REQUEST)
    instance = ser.save()
    return Response({'success': True, 'data': LessonSerializer(instance).data}, status=status.HTTP_201_CREATED)


@api_view(['GET'])
def get_lesson(request, pk):
    obj = Lesson.objects.filter(pk=pk).first()
    if not obj:
        return Response({'success': False, 'error': {'code': 'NOT_FOUND', 'message': '课程不存在'}}, status=status.HTTP_404_NOT_FOUND)
    return Response({'success': True, 'data': LessonSerializer(obj).data})


@api_view(['PATCH'])
def update_lesson(request, pk):
    obj = Lesson.objects.filter(pk=pk).first()
    if not obj:
        return Response({'success': False, 'error': {'code': 'NOT_FOUND', 'message': '课程不存在'}}, status=status.HTTP_404_NOT_FOUND)

    data = request.data

    # 预处理：weeks 数组 → 字符串
    weeks_val = data.get('weeks')
    if isinstance(weeks_val, list):
        data = {**data, 'weeks': ','.join(str(int(x)) for x in weeks_val if str(x).strip())}

    # 若传入 *_id，自动补齐名称字段，保持展示一致
    course_id = data.get('courseId') or data.get('course_id')
    if course_id:
        course = Course.objects.filter(pk=course_id).first()
        if course and not data.get('courseName') and not data.get('course_name'):
            data = {**data, 'courseName': course.name}

    teacher_id = data.get('teacherId') or data.get('teacher_id')
    if teacher_id:
        teacher = Teacher.objects.filter(pk=teacher_id).first()
        if teacher and not data.get('teacherName') and not data.get('teacher_name'):
            data = {**data, 'teacherName': teacher.name}

    class_id = data.get('classId') or data.get('class_id') or data.get('class_ref_id')
    if class_id:
        cls = SchoolClass.objects.filter(pk=class_id).first()
        if cls and not data.get('className') and not data.get('class_name'):
            data = {**data, 'className': cls.name}

    room_id = data.get('roomId') or data.get('room_id')
    if room_id:
        room = Room.objects.filter(pk=room_id).first()
        if room and not data.get('roomName') and not data.get('room_name'):
            data = {**data, 'roomName': room.name}

    ser = LessonSerializer(obj, data=data, partial=True)
    if not ser.is_valid():
        return Response({'success': False, 'error': {'code': 'VALIDATION_ERROR', 'message': '数据验证失败', 'details': ser.errors}}, status=status.HTTP_400_BAD_REQUEST)
    instance = ser.save()
    return Response({'success': True, 'data': LessonSerializer(instance).data})


@api_view(['DELETE'])
def delete_lesson(request, pk):
    obj = Lesson.objects.filter(pk=pk).first()
    if not obj:
        return Response({'success': False, 'error': {'code': 'NOT_FOUND', 'message': '课程不存在'}}, status=status.HTTP_404_NOT_FOUND)
    obj.delete()
    return Response({'success': True, 'data': {'deleted': True}}, status=status.HTTP_200_OK)


def _filter_by_week(qs, week: Optional[str]):
    if not week:
        return qs
    # 支持以下几种 weeks 格式：
    # - ""（空）：视为全周适用
    # - "1,3,5"（逗号分隔）
    # - "1-16"（范围，含端点）
    # 为保证兼容性与正确性，这里在 Python 层做一次过滤。
    try:
        week_int = int(str(week).strip())
    except Exception:
        week_int = None

    matched_ids: list = []
    for _id, w in qs.values_list('id', 'weeks'):
        s = (w or '').strip()
        if s == '':
            # 空 weeks 视为所有周次
            matched_ids.append(_id)
            continue
        # 逗号列表
        if ',' in s and '-' not in s:
            try:
                parts = [p.strip() for p in s.split(',') if p.strip()]
                if str(week) in parts:
                    matched_ids.append(_id)
                    continue
            except Exception:
                pass
        # 范围 a-b
        if '-' in s and week_int is not None:
            try:
                a_str, b_str = s.split('-', 1)
                a = int(a_str.strip())
                b = int(b_str.strip())
                if a <= week_int <= b:
                    matched_ids.append(_id)
                    continue
            except Exception:
                pass
        # 兜底：完全相等
        if s == str(week):
            matched_ids.append(_id)

    if not matched_ids:
        return qs.none()
    return qs.filter(id__in=matched_ids)


def _get_user_display_name(user) -> str:
    try:
        name = (getattr(user, 'get_full_name', lambda: '')() or '').strip()
    except Exception:
        name = ''
    if not name:
        # 退化为 first_name + last_name 或 username
        first = (getattr(user, 'first_name', '') or '').strip()
        last = (getattr(user, 'last_name', '') or '').strip()
        if first or last:
            name = f"{first}{last}".strip()
    if not name:
        name = (getattr(user, 'username', '') or '').strip()
    return name


def _apply_user_scope(request, qs):
    """按当前登录用户收敛可见课表范围：仅匹配教师姓名；学生不参与匹配。
    规则：用户显示名（全名/first+last/username）与教师表的 name 完全相同 → 仅看到该姓名的课表。
    未匹配到教师 → 不可见。
    """
    user = getattr(request, 'user', None)
    if not user or not getattr(user, 'is_authenticated', False):
        return qs.none()

    # 仅按姓名匹配教师
    display_name = _get_user_display_name(user)
    if not display_name:
        return qs.none()
    try:
        from apps.teachers.models import Teacher as T
        t = T.objects.filter(name=display_name).first()
    except Exception:
        t = None
    if not t:
        return qs.none()
    # 同时兼容按 teacher_id 关联与仅存储 teacher_name 的课次
    return qs.filter(Q(teacher_id=t.id) | Q(teacher_name=display_name))

@api_view(['GET'])
def school_timetable(request):
    term = request.query_params.get('term')
    week = request.query_params.get('week')
    qs = Lesson.objects.filter(term=term)
    qs = _apply_user_scope(request, qs)
    qs = _filter_by_week(qs, week)
    qs = qs.order_by('day_of_week','start_time','start_period')
    data = LessonSerializer(qs, many=True).data
    return Response({'success': True, 'data': {'lessons': data}})


@api_view(['GET'])
def class_timetable(request, pk):
    term = request.query_params.get('term')
    week = request.query_params.get('week')
    qs = Lesson.objects.filter(term=term, class_ref_id=pk)
    qs = _apply_user_scope(request, qs)
    qs = _filter_by_week(qs, week)
    qs = qs.order_by('day_of_week','start_time','start_period')
    data = LessonSerializer(qs, many=True).data
    return Response({'success': True, 'data': {'lessons': data}})


@api_view(['GET'])
def teacher_timetable(request, pk):
    term = request.query_params.get('term')
    week = request.query_params.get('week')
    qs = Lesson.objects.filter(term=term, teacher_id=pk)
    qs = _apply_user_scope(request, qs)
    qs = _filter_by_week(qs, week)
    qs = qs.order_by('day_of_week','start_time','start_period')
    data = LessonSerializer(qs, many=True).data
    return Response({'success': True, 'data': {'lessons': data}})


@api_view(['GET'])
def room_timetable(request, pk):
    term = request.query_params.get('term')
    week = request.query_params.get('week')
    qs = Lesson.objects.filter(term=term, room_id=pk)
    qs = _apply_user_scope(request, qs)
    qs = _filter_by_week(qs, week)
    qs = qs.order_by('day_of_week','start_time','start_period')
    data = LessonSerializer(qs, many=True).data
    return Response({'success': True, 'data': {'lessons': data}})


@api_view(['GET'])
def me_timetable(request):
    """根据当前登录用户角色返回个性化课表：
    - 教师用户：返回该教师的课表
    - 学生用户：返回其当前班级的课表
    - 管理员/其他：返回学校课表
    需要前端携带 JWT，后端通过 request.user 判定。
    支持查询参数：term, week
    """
    user = getattr(request, 'user', None)
    term = request.query_params.get('term')
    week = request.query_params.get('week')

    # 尝试定位教师
    teacher = None
    try:
        from apps.teachers.models import Teacher as T
        teacher = T.objects.filter(phone=getattr(user, 'username', None)).first() or T.objects.filter(email=getattr(user, 'email', None)).first()
    except Exception:
        teacher = None

    if teacher:
        qs = Lesson.objects.filter(term=term, teacher_id=teacher.id)
        qs = _filter_by_week(qs, week)
        qs = qs.order_by('day_of_week','start_time','start_period')
        return Response({'success': True, 'data': {'lessons': LessonSerializer(qs, many=True).data}})

    # 尝试定位学生（按用户名=学号 或 email/phone 匹配）
    student = None
    try:
        from apps.students.models import Student as S
        student = (
            S.objects.filter(student_id=getattr(user, 'username', None)).first()
            or S.objects.filter(email=getattr(user, 'email', None)).first()
            or S.objects.filter(phone=getattr(user, 'username', None)).first()
        )
    except Exception:
        student = None

    if student and student.current_class_id:
        qs = Lesson.objects.filter(term=term, class_ref_id=student.current_class_id)
        qs = _filter_by_week(qs, week)
        qs = qs.order_by('day_of_week','start_time','start_period')
        return Response({'success': True, 'data': {'lessons': LessonSerializer(qs, many=True).data}})

    # 默认：学校课表
    qs = Lesson.objects.filter(term=term)
    qs = _filter_by_week(qs, week)
    qs = qs.order_by('day_of_week','start_time','start_period')
    return Response({'success': True, 'data': {'lessons': LessonSerializer(qs, many=True).data}})

@api_view(['GET'])
def timetable_template(request):
    # 生成课程表样式 Excel 模板
    try:
        import openpyxl  # type: ignore
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    except Exception:
        return Response({'success': False, 'error': {'code':'DEPENDENCY_MISSING','message':'服务器缺少 openpyxl 依赖'}}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '课程表模板'

    days = ['星期一','星期二','星期三','星期四','星期五']
    periods = [
        '08:00-08:40','08:50-09:30','大课间 09:30-10:00','10:00-10:40','10:50-11:30','11:40-12:20',
        '午休 12:20-14:20','14:20-15:00','眼保健操 15:00-15:15','15:15-15:55','16:10-16:50','17:05-17:45'
    ]
    mode = request.query_params.get('mode', 'class')  # class: A列为班级；time: A列为时间

    # 列宽设置：第一列更宽显示时间
    ws.column_dimensions['A'].width = 18
    for col in range(2, 2 + 9 * len(days)):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 14

    # 顶部第一列标题
    ws.cell(row=1, column=1).value = '班级' if mode == 'class' else '时间/班级'
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')

    # 第一行：星期（每个星期占 9 列）
    col = 2
    for d in days:
        ws.cell(row=1, column=col).value = d
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+8)
        ws.cell(row=1, column=col).alignment = Alignment(horizontal='center', vertical='center')
        col += 9

    # 第二行：上午/下午（上午 1-5，下午 6-9；中间穿插特殊行如大课间、午休、眼保健操，仅作参考，可忽略）
    col = 2
    for _ in days:
        # 上午覆盖前5列
        ws.cell(row=2, column=col).value = '上午'
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col+4)
        ws.cell(row=2, column=col).alignment = Alignment(horizontal='center', vertical='center')
        # 下午覆盖后4列
        ws.cell(row=2, column=col+5).value = '下午'
        ws.merge_cells(start_row=2, start_column=col+5, end_row=2, end_column=col+8)
        ws.cell(row=2, column=col+5).alignment = Alignment(horizontal='center', vertical='center')
        col += 9

    # 第三行：节次序号 1..9 重复
    green = Font(color='008000', bold=True)
    row_idx = 3
    ws.cell(row=row_idx, column=1).value = ''
    col = 2
    for _ in days:
        for i in range(1, 10):
            c = ws.cell(row=row_idx, column=col)
            c.value = i
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.font = green
            col += 1
    # 数据区从第4行开始
    base_row = 4
    thin = Side(style='thin', color='000000')
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    if mode == 'class':
        # 从数据库读取班级名称（可选筛选：grade_id、grade_name、status）
        grade_id = request.query_params.get('grade_id')
        grade_name = request.query_params.get('grade_name')
        status_filter = request.query_params.get('status') or '在读'
        try:
            qs = SchoolClass.objects.all()
            if grade_id:
                qs = qs.filter(grade_id=grade_id)
            if grade_name:
                qs = qs.filter(grade__name=grade_name)
            if status_filter:
                qs = qs.filter(status=status_filter)
            class_names = list(qs.order_by('grade__name','name').values_list('name', flat=True))
        except Exception:
            class_names = []

        # 若数据库为空，给出占位示例
        if not class_names:
            class_names = ['一年级1班', '一年级2班', '一年级3班']

        for i, cls in enumerate(class_names):
            r = base_row + i
            ws.cell(row=r, column=1).value = cls
            ws.cell(row=r, column=1).alignment = Alignment(horizontal='center', vertical='center')
            for col in range(2, 2 + 9 * len(days)):
                ws.cell(row=r, column=col).border = border
                ws.cell(row=r, column=col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    else:
        # 时间行模板
        for i, p in enumerate(periods):
            r = base_row + i
            ws.cell(row=r, column=1).value = p
            ws.cell(row=r, column=1).alignment = Alignment(horizontal='center', vertical='center')
            for col in range(2, 2 + 9 * len(days)):
                ws.cell(row=r, column=col).border = border
                ws.cell(row=r, column=col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # 给表头加边框
    for r in range(1, base_row):
        for c in range(1, 2 + 9 * len(days)):
            ws.cell(row=r, column=c).border = border

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    resp = HttpResponse(bio.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="课程表导入模板.xlsx"'
    return resp


